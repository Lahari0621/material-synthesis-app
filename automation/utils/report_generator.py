from __future__ import annotations

import json
import shutil
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from config import settings


HEADERS = [
    "Test ID",
    "Module",
    "Test Name",
    "Priority",
    "Preconditions",
    "Test Steps",
    "Test Data",
    "Expected Result",
    "Actual Result",
    "Status",
    "Execution Time (ms)",
    "Failure Reason",
    "Screenshot",
    "Device Log",
]


def _load_results(input_root: Path) -> tuple[list[dict], list[dict]]:
    rows: dict[str, dict] = {}
    metadata = []
    for path in sorted(input_root.rglob("shard-*-results.json")):
        payload = json.loads(path.read_text(encoding="utf-8"))
        metadata.append(payload.get("metadata", {}))
        for row in payload.get("results", []):
            rows[row["test_id"]] = row
    return sorted(rows.values(), key=lambda row: row["test_id"]), metadata


def _metrics(rows: list[dict]) -> dict:
    counts = Counter(row["status"] for row in rows)
    executed = counts["PASSED"] + counts["FAILED"]
    pass_rate = counts["PASSED"] / executed * 100 if executed else 0.0
    fail_rate = counts["FAILED"] / executed * 100 if executed else 0.0
    critical = [
        row for row in rows
        if row["priority"] == "Critical" and row["status"] in ("PASSED", "FAILED")
    ]
    critical_failed = sum(row["status"] == "FAILED" for row in critical)
    critical_fail_rate = critical_failed / len(critical) if critical else 0.0
    return {
        "total": len(rows),
        "executed": executed,
        "passed": counts["PASSED"],
        "failed": counts["FAILED"],
        "skipped": counts["SKIPPED"],
        "blocked": counts["BLOCKED"],
        "pass_percentage": round(pass_rate, 2),
        "fail_percentage": round(fail_rate, 2),
        "critical_executed": len(critical),
        "critical_failed": critical_failed,
        "critical_fail_rate": round(critical_fail_rate, 4),
        "duration_seconds": round(
            sum(float(row.get("execution_time_ms", 0)) for row in rows) / 1000, 2
        ),
    }


def _style_header(sheet) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _append_rows(sheet, rows: list[dict]) -> None:
    sheet.append(HEADERS)
    for row in rows:
        sheet.append([
            row["test_id"],
            row["module"],
            row["test_name"],
            row["priority"],
            row["preconditions"],
            "\n".join(row["test_steps"]),
            json.dumps(row["test_data"], ensure_ascii=False),
            row["expected_result"],
            row.get("actual_result", ""),
            row["status"],
            row.get("execution_time_ms", 0),
            row.get("failure_reason", ""),
            row.get("screenshot", ""),
            row.get("device_log", ""),
        ])
    _style_header(sheet)
    for column in sheet.columns:
        letter = column[0].column_letter
        sheet.column_dimensions[letter].width = min(
            50, max(12, max(len(str(cell.value or "")) for cell in column) + 2)
        )


def _write_workbooks(rows: list[dict], metrics: dict, output: Path) -> None:
    excel = output / "Excel"
    excel.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    executed_sheet = wb.active
    executed_sheet.title = "Executed Test Cases"
    _append_rows(executed_sheet, rows)
    for status, name in (
        ("PASSED", "Passed Tests"),
        ("FAILED", "Failed Tests"),
        ("SKIPPED", "Skipped Tests"),
    ):
        _append_rows(
            wb.create_sheet(name),
            [row for row in rows if row["status"] == status],
        )

    metric_sheet = wb.create_sheet("Execution Metrics")
    metric_sheet.append(["Metric", "Value"])
    for key, value in metrics.items():
        metric_sheet.append([key.replace("_", " ").title(), value])
    _style_header(metric_sheet)

    defects = wb.create_sheet("Defect Summary")
    defects.append(["Test ID", "Module", "Priority", "Failure Reason"])
    for row in rows:
        if row["status"] == "FAILED":
            defects.append([
                row["test_id"], row["module"], row["priority"],
                row.get("failure_reason", ""),
            ])
    _style_header(defects)

    rates = wb.create_sheet("Pass Rate Summary")
    rates.append(["Module", "Total", "Passed", "Failed", "Blocked", "Pass Rate"])
    grouped = defaultdict(list)
    for row in rows:
        grouped[row["module"]].append(row)
    for module, module_rows in sorted(grouped.items()):
        counter = Counter(row["status"] for row in module_rows)
        executed = counter["PASSED"] + counter["FAILED"]
        rate = counter["PASSED"] / executed * 100 if executed else 0
        rates.append([
            module, len(module_rows), counter["PASSED"], counter["FAILED"],
            counter["BLOCKED"], round(rate, 2),
        ])
    _style_header(rates)
    wb.save(excel / "Automation_Test_Report.xlsx")

    for filename, selected in (
        ("Passed_Test_Cases.xlsx", [r for r in rows if r["status"] == "PASSED"]),
        ("Failed_Test_Cases.xlsx", [r for r in rows if r["status"] == "FAILED"]),
    ):
        book = Workbook()
        _append_rows(book.active, selected)
        book.active.title = filename.removesuffix(".xlsx")[:31]
        book.save(excel / filename)

    summary = Workbook()
    sheet = summary.active
    sheet.title = "Execution Summary"
    sheet.append(["Metric", "Value"])
    for key, value in metrics.items():
        sheet.append([key.replace("_", " ").title(), value])
    _style_header(sheet)
    summary.save(excel / "Execution_Summary.xlsx")


def _write_html(
    rows: list[dict], metrics: dict, metadata: list[dict], output: Path
) -> None:
    html_dir = output / "HTML"
    html_dir.mkdir(parents=True, exist_ok=True)
    device = metadata[0] if metadata else {}
    failures = [row for row in rows if row["status"] == "FAILED"]
    failure_gallery = "\n".join(
        f"<section><h3>{row['test_id']} · {row['test_name']}</h3>"
        f"<pre>{row.get('failure_reason', '')[:3000]}</pre>"
        + (
            f"<img style='max-width:360px' src='../Screenshots/"
            f"{Path(row['screenshot']).name}' alt='{row['test_id']} screenshot'>"
            if row.get("screenshot")
            else "<p>No screenshot was available.</p>"
        )
        + "</section>"
        for row in failures
    )
    table_rows = "\n".join(
        f"<tr class='{row['status'].lower()}'><td>{row['test_id']}</td>"
        f"<td>{row['module']}</td><td>{row['test_name']}</td>"
        f"<td>{row['priority']}</td><td>{row['status']}</td>"
        f"<td>{row.get('execution_time_ms', 0)}</td>"
        f"<td><pre>{row.get('failure_reason', '')[:1000]}</pre></td></tr>"
        for row in rows
    )
    html = f"""<!doctype html><html><head><meta charset="utf-8">
<title>Android Appium E2E Report</title>
<style>
body{{font-family:Arial;margin:24px;background:#f4f7fb;color:#17233b}}
.cards{{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:12px}}
.card{{background:white;padding:18px;border-radius:10px;box-shadow:0 2px 8px #ccd}}
.value{{font-size:28px;font-weight:bold}} table{{width:100%;border-collapse:collapse;background:white}}
th,td{{padding:8px;border:1px solid #ddd;text-align:left;vertical-align:top}}
th{{background:#1f4e79;color:white}} .passed{{background:#e8f5e9}}
.failed{{background:#ffebee}} .blocked{{background:#fff8e1}} pre{{white-space:pre-wrap}}
</style></head><body>
<h1>Android Appium E2E Execution Report</h1>
<p>Build {settings.BUILD_NUMBER} · Commit {settings.COMMIT_SHA[:8]} ·
Branch {settings.BRANCH} · Device {device.get('device', settings.DEVICE_NAME)} ·
Android {device.get('android_version', settings.PLATFORM_VERSION)} ·
App Version {settings.APP_VERSION}</p>
<div class="cards">
{''.join(f'<div class="card"><div>{k.replace("_"," ").title()}</div><div class="value">{v}</div></div>' for k,v in metrics.items() if k in ("total","executed","passed","failed","skipped","blocked","pass_percentage","duration_seconds"))}
</div>
<h2>Test Cases</h2><table><thead><tr><th>ID</th><th>Module</th><th>Name</th>
<th>Priority</th><th>Status</th><th>Time ms</th><th>Failure</th></tr></thead>
<tbody>{table_rows}</tbody></table>
<h2>Failure Details ({len(failures)})</h2>
{failure_gallery or '<p>No failures.</p>'}
<h2>Historical Trends</h2>
<p><a href="trends.html">Open historical trend index</a>. Archived runs are
published under <code>reports/history/build-N/</code>.</p>
</body></html>"""
    (html_dir / "execution-report.html").write_text(html, encoding="utf-8")

    dashboard = html.replace(
        "<h2>Test Cases</h2>", "<h2>Dashboard</h2><p>Module-level quality metrics are in the Excel report.</p><h2>Test Cases</h2>"
    )
    (html_dir / "dashboard.html").write_text(dashboard, encoding="utf-8")

    trends = """<!doctype html><html><head><meta charset="utf-8"><title>Trends</title>
<style>body{font-family:Arial;margin:30px}</style></head><body>
<h1>Historical Trends</h1><p>Each GitHub Pages deployment is archived under
<code>reports/history/build-N/</code>. Compare archived execution summaries for trends.</p>
</body></html>"""
    (html_dir / "trends.html").write_text(trends, encoding="utf-8")


def _write_summary(rows: list[dict], metrics: dict, output: Path) -> None:
    passed = [row for row in rows if row["status"] == "PASSED"]
    failed = [row for row in rows if row["status"] == "FAILED"]
    skipped = [
        row for row in rows if row["status"] in ("SKIPPED", "BLOCKED")
    ]
    lines = [
        "# Android Appium E2E Execution Summary",
        "",
        f"Build Number: {settings.BUILD_NUMBER}",
        f"Execution Date: {datetime.now(timezone.utc).isoformat()}",
        f"Git Commit: {settings.COMMIT_SHA}",
        f"Branch: {settings.BRANCH}",
        "",
        f"APK Version: {settings.APP_VERSION}",
        "",
        f"Device: {settings.DEVICE_NAME}",
        f"Android Version: {settings.PLATFORM_VERSION}",
        "",
        "## Execution Metrics",
        "",
        f"- Total Test Cases: {metrics['total']}",
        f"- Executed: {metrics['executed']}",
        f"- Passed: {metrics['passed']}",
        f"- Failed: {metrics['failed']}",
        f"- Skipped: {metrics['skipped']}",
        f"- Blocked: {metrics['blocked']}",
        f"- Pass Percentage: {metrics['pass_percentage']}%",
        f"- Fail Percentage: {metrics['fail_percentage']}%",
        f"- Execution Duration: {metrics['duration_seconds']} seconds",
        "",
        "## PASSED TESTS",
        *[f"✓ {r['test_id']} - {r['test_name']}" for r in passed],
        "",
        "## FAILED TESTS",
        *[
            f"✗ {r['test_id']} - {r['test_name']}\n  Reason: {r.get('failure_reason','')[:500]}"
            for r in failed
        ],
        "",
        "## SKIPPED / BLOCKED TESTS",
        *[
            f"- {r['test_id']} - {r['test_name']}\n  Reason: {r.get('failure_reason','')}"
            for r in skipped
        ],
    ]
    (output / "Summary" / "summary.md").write_text(
        "\n".join(lines), encoding="utf-8"
    )


def generate(input_root: Path, output: Path) -> tuple[dict, int]:
    rows, metadata = _load_results(input_root)
    if len(rows) != 510:
        raise RuntimeError(
            f"Expected reports for 510 collected cases, found {len(rows)}"
        )
    metrics = _metrics(rows)
    output.mkdir(parents=True, exist_ok=True)
    (output / "JSON").mkdir(parents=True, exist_ok=True)
    (output / "Screenshots").mkdir(parents=True, exist_ok=True)
    (output / "Logs").mkdir(parents=True, exist_ok=True)
    (output / "Summary").mkdir(parents=True, exist_ok=True)
    (output / "JSON" / "execution-results.json").write_text(
        json.dumps(
            {"metadata": metadata, "metrics": metrics, "results": rows},
            indent=2,
        ),
        encoding="utf-8",
    )
    _write_workbooks(rows, metrics, output)
    _write_html(rows, metrics, metadata, output)
    _write_summary(rows, metrics, output)

    for pattern, destination in (
        ("**/*.png", output / "Screenshots"),
        ("**/*log*.txt", output / "Logs"),
        ("**/appium*.log", output / "Logs"),
    ):
        for source in input_root.glob(pattern):
            if source.is_file():
                shutil.copy2(source, destination / source.name)

    pass_ok = metrics["pass_percentage"] >= 95.0
    critical_ok = (
        metrics["critical_fail_rate"]
        <= settings.CRITICAL_FAILURE_THRESHOLD
    )
    return metrics, 0 if pass_ok and critical_ok else 1


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, default=settings.RESULTS_ROOT)
    parser.add_argument("--output", type=Path, default=settings.RESULTS_ROOT)
    args = parser.parse_args()
    generated_metrics, code = generate(args.input, args.output)
    print(json.dumps(generated_metrics, indent=2))
    raise SystemExit(code)

