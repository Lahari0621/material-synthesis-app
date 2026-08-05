"""Excel analysis reporter for Appium E2E results."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


class ExcelReporter:
    def __init__(self, output_path: Path):
        self.output_path = Path(output_path)
        self.rows: list[dict] = []

    def add_result(
        self,
        test_id: str,
        module: str,
        scenario: str,
        status: str,
        duration_ms: float,
        details: str = "",
    ) -> None:
        self.rows.append(
            {
                "test_id": test_id,
                "module": module,
                "scenario": scenario,
                "status": status.upper(),
                "duration_ms": round(duration_ms, 1),
                "details": details,
                "executed_at": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC"),
            }
        )

    def write(self) -> Path:
        wb = Workbook()

        summary = wb.active
        summary.title = "Summary"
        passed = sum(1 for row in self.rows if row["status"] == "PASSED")
        failed = sum(1 for row in self.rows if row["status"] == "FAILED")
        total = len(self.rows)
        pass_rate = (passed / total * 100.0) if total else 0.0

        summary_headers = ["Metric", "Value"]
        summary.append(summary_headers)
        summary.append(["Total Test Cases", total])
        summary.append(["Passed", passed])
        summary.append(["Failed", failed])
        summary.append(["Pass Rate (%)", round(pass_rate, 2)])
        summary.append(["Generated At", datetime.utcnow().isoformat() + "Z"])
        summary.append(["Framework", "Appium Android E2E + Excel Analysis"])
        self._style_header(summary)

        details = wb.create_sheet("Test Cases")
        details.append(
            [
                "Test ID",
                "Module",
                "Scenario",
                "Status",
                "Duration (ms)",
                "Details",
                "Executed At",
            ]
        )
        for row in self.rows:
            details.append(
                [
                    row["test_id"],
                    row["module"],
                    row["scenario"],
                    row["status"],
                    row["duration_ms"],
                    row["details"],
                    row["executed_at"],
                ]
            )
        self._style_header(details)
        self._color_status(details)

        analysis = wb.create_sheet("Excel Analysis")
        analysis.append(["Module", "Total", "Passed", "Failed", "Pass Rate (%)"])
        modules: dict[str, dict[str, int]] = {}
        for row in self.rows:
            bucket = modules.setdefault(
                row["module"], {"total": 0, "passed": 0, "failed": 0}
            )
            bucket["total"] += 1
            if row["status"] == "PASSED":
                bucket["passed"] += 1
            else:
                bucket["failed"] += 1
        for module, stats in sorted(modules.items()):
            rate = (
                stats["passed"] / stats["total"] * 100.0 if stats["total"] else 0.0
            )
            analysis.append(
                [
                    module,
                    stats["total"],
                    stats["passed"],
                    stats["failed"],
                    round(rate, 2),
                ]
            )
        self._style_header(analysis)

        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(self.output_path)
        return self.output_path

    @staticmethod
    def _style_header(sheet) -> None:
        fill = PatternFill("solid", fgColor="1F4E79")
        font = Font(color="FFFFFF", bold=True)
        for cell in sheet[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center")

    @staticmethod
    def _color_status(sheet) -> None:
        green = PatternFill("solid", fgColor="C6EFCE")
        red = PatternFill("solid", fgColor="FFC7CE")
        for row in sheet.iter_rows(min_row=2, min_col=4, max_col=4):
            cell = row[0]
            if str(cell.value).upper() == "PASSED":
                cell.fill = green
            elif str(cell.value).upper() == "FAILED":
                cell.fill = red
